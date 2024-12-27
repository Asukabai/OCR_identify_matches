import requests
import json
import openpyxl
import os
import base64
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from PIL import Image
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
import fitz  # PyMuPDF

# compress_image 功能：压缩图片，确保其文件大小不超过指定的最大值。 参数：image_path (str): 图片文件的路径。 max_size_kb (int, 默认值 1024): 图片的最大文件大小（单位：KB）。
# 逻辑：
#   打开图片并获取其 EXIF 信息。
#   获取原始图片的尺寸。
#   使用循环尝试不同的质量参数，直到文件大小小于或等于 max_size_kb 或质量参数降到 10 以下。
#   保存压缩后的图片，确保其尺寸与原始图片一致。
#   返回压缩后的图片对象。

def compress_image(image_path, max_size_kb=1024):
    print("进入压缩方法")
    # 打开图片
    img = Image.open(image_path)
    # 保留图片的 EXIF 信息
    exif_data = img.info.get('exif', None)
    # 获取原始图片的尺寸
    original_size = img.size
    # 尝试不同的质量参数，直到文件大小小于或等于 max_size_kb
    quality = 95
    while True:
        # 保存压缩后的图片，保留 EXIF 信息
        img.save(image_path, format=img.format, optimize=True, quality=quality, exif=exif_data)
        # 检查文件大小
        file_size_kb = os.path.getsize(image_path) / 1024
        if file_size_kb <= max_size_kb or quality <= 10:
            break
        # 降低质量参数
        quality -= 5
    # 确保压缩后的图片尺寸与原始图片一致
    img = Image.open(image_path)
    if img.size != original_size:
        img = img.resize(original_size)
        img.save(image_path, format=img.format, optimize=True, quality=quality, exif=exif_data)
    print(f"图片压缩完成，文件大小: {file_size_kb:.2f} KB")
    return img  # 返回压缩后的图片对象


# post_url 功能：发送 POST 请求到 OCR API，并将结果写入 Excel 文件。 参数： base64_strings (list): 图片的 Base64 编码字符串列表。excel_file_path (str): Excel 文件的路径。file_name (str): 文件名。imgs (list): 图片对象列表。
# 逻辑：
#   构建请求数据，包括 Base64 编码的图片字符串和选项。
#   设置请求头和重试策略。
#   发送 POST 请求到指定 URL。
#   解析响应数据并将其写入 Excel 文件的相应单元格。
#   调用 insert_images_into_excel 方法将图片插入到 Excel 文件中。
#   保存 Excel 文件。
#   处理请求异常并打印错误信息。


def post_url(base64_strings, excel_file_path, file_name, imgs):
    url = "http://127.0.0.1:1224/api/ocr"
    data = {
        "base64": base64_strings[0],
        "options": {
            "data.format": "text",
        }
    }
    headers = {"Content-Type": "application/json"}
    data_str = json.dumps(data)
    session = requests.Session()
    retries = Retry(total=4, backoff_factor=2, status_forcelist=[502, 503, 504])
    session.mount('http://', HTTPAdapter(max_retries=retries))
    try:
        response = session.post(url, data=data_str, headers=headers, timeout=100)
        response.raise_for_status()
        res_dict = json.loads(response.text)
        # 写入 Excel
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet = workbook['Sheet']
        row = sheet.max_row + 1
        base_name = os.path.splitext(file_name)[0]  # 去掉文件后缀
        sheet[f'A{row}'] = base_name  # 文件名放在第一列
        sheet[f'B{row}'] = res_dict.get('data', '')  # 文件内容放在第二列

        # 插入图片
        insert_images_into_excel(sheet, row, imgs)

        # 保存工作簿
        workbook.save(excel_file_path)
    except requests.exceptions.RequestException as e:
        print(f"请求失败: {e}")
    print(f'post_url 方法调用完成！文件: {file_name}')


# insert_images_into_excel 功能：将图片插入到 Excel 工作表中。 参数： sheet (openpyxl.worksheet.worksheet.Worksheet): Excel 工作表对象。 row (int): 插入图片的行号。 imgs (list): 图片对象列表。
# 逻辑：
#    初始化列索引为 3（即 C 列）。
#    遍历图片对象列表，生成正确的列名。
#    将图片对象插入到指定的单元格。
#    更新列索引以插入下一张图片。

def insert_images_into_excel(sheet, row, imgs):
    col = 3  # 从 C 列开始
    for img in imgs:
        # 生成正确的列名
        col_letter = openpyxl.utils.get_column_letter(col)
        img_path = f"{col_letter}{row}"
        img_obj = ExcelImage(img)
        sheet.add_image(img_obj, img_path)
        col += 1

# fix_pic_file 功能：处理图片文件，包括压缩和 Base64 编码。 参数：image_path (str): 图片文件的路径。
# 逻辑：
#    判断文件大小，如果超过 1MB，则调用 compress_image 方法进行压缩。
#    打开图片文件并读取其内容，进行 Base64 编码。
#    去掉 Base64 编码头部（如果存在）。
#    返回 Base64 编码字符串和图片对象。

def fix_pic_file(image_path):
    # 判断文件大小
    file_size_kb = os.path.getsize(image_path) / 1024
    if file_size_kb > 1024:  # 1MB
        # 压缩图片
        img = compress_image(image_path)
    else:
        img = Image.open(image_path)
    with open(image_path, "rb") as image_file:
        encoded_string = base64.b64encode(image_file.read()).decode('utf-8')

    # 去掉 base64 编码头部
    if encoded_string.startswith('data:image/png;base64,'):
        encoded_string = encoded_string.replace('data:image/png;base64,', '')
    elif encoded_string.startswith('data:image/jpg;base64,'):
        encoded_string = encoded_string.replace('data:image/jpg;base64,', '')

    print(f'fix_pic_file 方法调用完成！文件: {os.path.basename(image_path)}')
    return encoded_string, img



# 功能：将 PDF 文件转换为图片文件。 参数：pdf_path (str): PDF 文件的路径   ;  output_folder (str): 输出图片文件的目录。
# 逻辑：
#     打开 PDF 文件。
#     遍历每一页，将其转换为图片并保存到指定目录。
#     返回图片文件路径列表和 PDF 文件的基本名称。

def convert_pdf_to_image(pdf_path, output_folder):
    # 打开 PDF 文件
    pdf_document = fitz.open(pdf_path)
    base_name = os.path.splitext(os.path.basename(pdf_path))[0]
    images = []
    for page_num in range(len(pdf_document)):
        page = pdf_document.load_page(page_num)
        pix = page.get_pixmap()
        img_path = os.path.join(output_folder, f"{base_name}_page_{page_num + 1}.png")
        pix.save(img_path)
        images.append(img_path)
    return images, base_name


# 功能：批量处理指定文件夹中的图片和 PDF 文件。 参数：folder_path (str): 包含图片和 PDF 文件的文件夹路径  ;  excel_file_path (str): Excel 文件的路径。

# 逻辑：获取文件夹中所有符合条件的文件（PNG、JPG、JPEG、PDF）。
# 遍历每个文件：
# 如果是 PDF 文件，调用 convert_pdf_to_image 方法将其转换为图片。
#    对每张图片调用 fix_pic_file 方法进行处理，获取 Base64 编码字符串和图片对象。
#    调用 post_url 方法将处理结果写入 Excel 文件。
# 如果是图片文件，直接调用 fix_pic_file 和 post_url 方法进行处理。


def batch_process(folder_path, excel_file_path):
    files = [f for f in os.listdir(folder_path) if f.endswith(('.png', '.jpg', '.jpeg', '.pdf'))]
    for file in files:
        file_path = os.path.join(folder_path, file)
        if file.endswith('.pdf'):
            # 转换 PDF 为图片
            images, base_name = convert_pdf_to_image(file_path, folder_path)
            base64_strings = []
            compressed_imgs = []
            for img_path in images:
                base64_string, img = fix_pic_file(img_path)
                base64_strings.append(base64_string)
                compressed_imgs.append(img)
            post_url(base64_strings, excel_file_path, base_name, compressed_imgs)
        else:
            base64_string, img = fix_pic_file(file_path)
            post_url([base64_string], excel_file_path, file, [img])




# 批量处理指定文件夹中的 图片和 PDF 文件，通过 OCR API 获取文本内容，并将结果写入 Excel 文件中（同时，处理过程中会对图片进行压缩和 Base64 编码)

# 1、指定文件夹路径和 Excel 文件路径。
# 2、如果 Excel 文件不存在，创建一个新的 Excel 文件。
# 3、调用 batch_process 方法批量处理文件夹中的图片和 PDF 文件。
# 4、打印处理结束信息。

if __name__ == '__main__':
    # 指定文件夹路径和 Excel 文件路径
    folder_path = 'D:\\ssproject\\ssprice\\采购合同\\pic_1 - 副本'
    excel_file_path = 'D:\\ssproject\\ssprice\\采购合同\\pic_1 - 副本\\excel.xlsx'
    # 创建 Excel 文件（如果不存在）
    if not os.path.exists(excel_file_path):
        workbook = Workbook()
        workbook.save(excel_file_path)
    # 批量处理图片
    batch_process(folder_path, excel_file_path)
    print('处理结束！')




